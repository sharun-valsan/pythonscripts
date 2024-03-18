#!/usr/bin/env python3
import csv
from openpyxl import load_workbook
import requests
import os

def download_data_file(url, output_file):
    response = requests.get(url)
    if response.status_code == 200:
        with open(output_file, 'wb') as f:
            f.write(response.content)
    else:
        print("Failed to download the data file from the provided URL.")

def update_docs_data(data_file, docs_file):
    # Step 1: Read data from the data file
    with open(data_file, 'r') as data_file:
        reader = csv.reader(data_file, delimiter=';')
        data = list(reader)

    # Step 2: Read strings from the 8th row of the docs file and update associated 9th row
    workbook = load_workbook(filename=docs_file)
    sheet = workbook.active
    for row_number in range(1, sheet.max_row + 1):
        eighth_row_value = sheet.cell(row=row_number, column=8).value
        for row in data:
            if eighth_row_value in row:
                ninth_row_value = row[4]  # Assuming index starts from 0
                sheet.cell(row=row_number, column=9).value = ninth_row_value
                break  # Assuming only one match is expected, exit loop after update
    
    # Save the changes to the Excel file
    workbook.save(docs_file)
    workbook.close()

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.realpath(__file__))
    data_file_path = os.path.join(script_dir, "NAVAll.txt")

    data_file_url = "https://www.amfiindia.com/spages/NAVAll.txt"
    download_data_file(data_file_url, data_file_path)

    docs_file = os.path.join(script_dir, "mf_worksheet.xlsx")

    update_docs_data(data_file_path, docs_file)

    # Clean up: Delete the downloaded file after use
    #os.remove(data_file_path)
